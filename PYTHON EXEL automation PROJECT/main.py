import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)



values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=3, max_col=3)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions_graph.xlsx')